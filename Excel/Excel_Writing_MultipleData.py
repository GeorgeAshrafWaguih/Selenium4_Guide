import openpyxl

# Writing Multiple Data
file = '/data1.xlsx'
workbook = openpyxl.load_workbook(file)  # load the workbook
sheet = workbook.active
# workbook.active (In case to write in the current sheet

sheet.cell(1, 1).value = 1
sheet.cell(1, 2).value = "Smith"
sheet.cell(1, 3).value = "Engineer"

sheet.cell(2, 1).value = 2
sheet.cell(2, 2).value = "John"
sheet.cell(2, 3).value = "Doctor"

sheet.cell(3, 1).value = 3
sheet.cell(3, 2).value = "David"
sheet.cell(1, 3).value = "Pilot"

workbook.save(file)  # save file 
