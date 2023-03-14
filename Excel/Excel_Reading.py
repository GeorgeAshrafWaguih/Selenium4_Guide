import openpyxl

# File -> Workbook -> Sheet -> Rows -> Cells
# Path of the execl file
file = '/data.xlsx'
workbook = openpyxl.load_workbook(file)
sheet = workbook["Sheet1"]

# Find no of rows and columns in the sheet
rows = sheet.max_row  # count number of rows
columns = sheet.max_column  # count number of columns

# Reading all the rows and columns from the sheet
for r in range(1, rows+1):
    for c in range(1, columns+1):
        print(sheet.cell(r, c).value, end="    ")
    print()
