import openpyxl
from openpyxl.styles import PatternFill


def getRowCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_row


def getColumnCount(file, sheetname):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.max_column


def readData(file, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    return sheet.cell(rownum, columnum).value


def writeData(file, sheetname, rownum, columnum, data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    sheet.cell(rownum, columnum).value = data
    workbook.save(file)


def fillGreenColor(file, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    greenfill = PatternFill(start_color='60b212',
                            end_color='60b212',
                            fill_type='solid')
    sheet.cell(rownum, columnum).fill = greenfill
    workbook.save(file)


def fillRedColor(file, sheetname, rownum, columnum):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook[sheetname]
    redfill = PatternFill(start_color='ff0000',
                            end_color='ff0000',
                            fill_type='solid')
    sheet.cell(rownum, columnum).fill = redfill
    workbook.save(file)
